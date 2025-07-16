#!/usr/bin/env python3
import pandas as pd
import numpy as np
import openpyxl

def analyze_excel_file(file_path):
    print("="*50)
    print("ANALYZING EXCEL FILE: tinhtoan.xlsx")
    print("="*50)
    
    try:
        # Read with openpyxl to get more detailed information
        wb = openpyxl.load_workbook(file_path)
        print(f"Sheet names: {wb.sheetnames}")
        
        # Analyze each sheet
        for sheet_name in wb.sheetnames:
            print(f"\n--- Sheet: {sheet_name} ---")
            ws = wb[sheet_name]
            
            # Get sheet dimensions
            max_row = ws.max_row
            max_col = ws.max_column
            print(f"Dimensions: {max_row} rows x {max_col} columns")
            
            # Read the sheet with pandas for better analysis
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            print(f"DataFrame shape: {df.shape}")
            
            # Display all content
            print("\nAll content:")
            for i in range(min(50, len(df))):  # Show first 50 rows
                row_content = []
                for j in range(min(10, len(df.columns))):  # Show first 10 columns
                    cell_value = df.iloc[i, j]
                    if pd.isna(cell_value):
                        row_content.append("")
                    else:
                        row_content.append(str(cell_value))
                print(f"Row {i+1}: {row_content}")
            
            # Look for specific patterns related to O calculation
            print("\n--- Looking for O calculation patterns ---")
            for i in range(len(df)):
                for j in range(len(df.columns)):
                    cell_value = str(df.iloc[i, j]).strip().lower()
                    if any(keyword in cell_value for keyword in ['output', 'calculating o', 'o =', 'o(', 'when', 'parameter', '= 0']):
                        print(f"Found at Row {i+1}, Col {j+1}: {df.iloc[i, j]}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error with openpyxl: {e}")
        
        # Fallback to pandas only
        try:
            xl_file = pd.ExcelFile(file_path)
            print(f"Sheet names (pandas): {xl_file.sheet_names}")
            
            for sheet_name in xl_file.sheet_names:
                print(f"\n--- Sheet: {sheet_name} (pandas) ---")
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                print(f"Shape: {df.shape}")
                
                # Display content
                print("\nContent:")
                for i in range(min(30, len(df))):
                    row_data = []
                    for j in range(min(8, len(df.columns))):
                        cell_value = df.iloc[i, j]
                        if pd.isna(cell_value):
                            row_data.append("")
                        else:
                            row_data.append(str(cell_value))
                    print(f"Row {i+1}: {row_data}")
                
        except Exception as e2:
            print(f"Error with pandas: {e2}")

if __name__ == "__main__":
    file_path = "/home/binh/Documents/excel2site/tinhtoan.xlsx"
    analyze_excel_file(file_path)